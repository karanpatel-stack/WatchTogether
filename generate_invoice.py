import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Invoice"

# ── Colors ──
DARK_BG = "1A1A2E"
ACCENT = "6C5CE7"
ACCENT_LIGHT = "A29BFE"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
MED_GRAY = "E9ECEF"
TEXT_DARK = "2D3436"
TEXT_MED = "636E72"
SECTION_BG = "EDE7F6"
SUBTOTAL_BG = "F3E5F5"
TOTAL_BG = "1A1A2E"

# ── Fonts ──
title_font = Font(name="Calibri", size=28, bold=True, color=WHITE)
subtitle_font = Font(name="Calibri", size=11, color=ACCENT_LIGHT)
header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
section_font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
item_font = Font(name="Calibri", size=10, color=TEXT_DARK)
item_desc_font = Font(name="Calibri", size=9, color=TEXT_MED, italic=True)
subtotal_font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
total_font = Font(name="Calibri", size=14, bold=True, color=WHITE)
total_amount_font = Font(name="Calibri", size=18, bold=True, color=WHITE)
label_font = Font(name="Calibri", size=10, color=TEXT_MED)
value_font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
notes_font = Font(name="Calibri", size=9, color=TEXT_MED)

# ── Borders ──
thin_border = Border(
    bottom=Side(style="thin", color=MED_GRAY)
)
header_border = Border(
    bottom=Side(style="medium", color=ACCENT)
)

# ── Column widths ──
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 48
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 3

# ── Helper functions ──
def merge_fill(row, col_start, col_end, value, font, fill=None, alignment=None, border=None):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = font
    if fill:
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).border = border
    return cell

def set_row_height(row, height):
    ws.row_dimensions[row].height = height

# ══════════════════════════════════════════════════════════════
# HEADER SECTION
# ══════════════════════════════════════════════════════════════
dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")

for r in range(1, 7):
    set_row_height(r, 6 if r == 1 else None)
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = dark_fill

set_row_height(2, 45)
merge_fill(2, 2, 4, "INVOICE", title_font, dark_fill, Alignment(vertical="center"))

# Invoice number and date on the right
ws.cell(row=2, column=5, value="Invoice #").font = Font(name="Calibri", size=9, color=ACCENT_LIGHT)
ws.cell(row=2, column=5).fill = dark_fill
ws.cell(row=2, column=5).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=2, column=6, value="WT-2026-002").font = Font(name="Calibri", size=12, bold=True, color=WHITE)
ws.cell(row=2, column=6).fill = dark_fill
ws.cell(row=2, column=6).alignment = Alignment(vertical="center")

set_row_height(3, 20)
merge_fill(3, 2, 4, "Full-Stack Web Application Development", subtitle_font, dark_fill, Alignment(vertical="center"))

ws.cell(row=3, column=5, value="Date:").font = Font(name="Calibri", size=9, color=ACCENT_LIGHT)
ws.cell(row=3, column=5).fill = dark_fill
ws.cell(row=3, column=5).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=3, column=6, value=date.today().strftime("%B %d, %Y")).font = Font(name="Calibri", size=10, color=WHITE)
ws.cell(row=3, column=6).fill = dark_fill
ws.cell(row=3, column=6).alignment = Alignment(vertical="center")

set_row_height(4, 5)
set_row_height(5, 22)
merge_fill(5, 2, 6, "WatchTogether — Self-Hosted YouTube Watch Party Platform", Font(name="Calibri", size=10, color=ACCENT_LIGHT, italic=True), dark_fill, Alignment(vertical="center"))

set_row_height(6, 6)

# ══════════════════════════════════════════════════════════════
# PROJECT SUMMARY BAR
# ══════════════════════════════════════════════════════════════
r = 7
set_row_height(r, 8)

r = 8
accent_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
set_row_height(r, 32)
labels = ["TECH STACK", "CODEBASE", "FEATURES", "RATE"]
values = ["React + Node.js + Socket.IO + WebRTC", "4,800+ lines across 35+ files", "12 major systems", "$250.00/hr"]
for i, (lab, val) in enumerate(zip(labels, values)):
    col = 2 + i
    if i < 3:
        ws.cell(row=r, column=col+1, value=f"{lab}: {val}").font = Font(name="Calibri", size=9, bold=True, color=WHITE)
        ws.cell(row=r, column=col+1).fill = accent_fill
        ws.cell(row=r, column=col+1).alignment = Alignment(vertical="center", horizontal="center")

# Simpler approach - merge the accent bar
for c in range(2, 7):
    ws.cell(row=r, column=c).fill = accent_fill

ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(row=r, column=2, value="React + Node.js + Socket.IO + mediasoup  |  5,800+ lines  |  36+ files  |  13 major systems").font = Font(name="Calibri", size=9, bold=True, color=WHITE)
ws.cell(row=r, column=2).alignment = Alignment(vertical="center", horizontal="center")
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
ws.cell(row=r, column=4, value="Rate: $250.00 / hour").font = Font(name="Calibri", size=10, bold=True, color=WHITE)
ws.cell(row=r, column=4).alignment = Alignment(vertical="center", horizontal="center")

r = 9
set_row_height(r, 10)

# ══════════════════════════════════════════════════════════════
# TABLE HEADER
# ══════════════════════════════════════════════════════════════
r = 10
set_row_height(r, 30)
headers = ["", "Task", "Description", "Hours", "Rate", "Amount"]
accent_dark = PatternFill(start_color="2D2856", end_color="2D2856", fill_type="solid")
for i, h in enumerate(headers):
    col = i + 1
    cell = ws.cell(row=r, column=col, value=h)
    cell.font = header_font
    cell.fill = accent_dark
    cell.alignment = Alignment(vertical="center", horizontal="left" if i < 3 else "center")
    cell.border = header_border
ws.cell(row=r, column=5).alignment = Alignment(vertical="center", horizontal="center")
ws.cell(row=r, column=6).alignment = Alignment(vertical="center", horizontal="right")
ws.cell(row=r, column=7).fill = accent_dark

# ══════════════════════════════════════════════════════════════
# LINE ITEMS DATA
# ══════════════════════════════════════════════════════════════
RATE = 250.00

sections = [
    {
        "name": "1. PROJECT ARCHITECTURE & SETUP",
        "items": [
            ("Initialize React + Vite + TypeScript frontend", "Project scaffolding, directory structure, ESLint, Vite config with dev proxy, TypeScript strict mode config", 3),
            ("Initialize Express + Socket.IO server", "TypeScript server setup, typed Socket.IO events (29 event types), CORS config, health check endpoint", 3),
            ("Docker Compose multi-service stack", "Three-service orchestration (server, frontend, tunnel), Dockerfiles with multi-stage builds, network config", 2.5),
            ("Nginx configuration", "SPA fallback routing, WebSocket upgrade headers (proxy_read_timeout 86400), API reverse proxy, static asset caching (1yr), gzip compression", 2),
            ("Tailwind CSS & PostCSS pipeline", "Custom accent color system via CSS variables with opacity support, custom animations (fadeIn, slideUp, pulse-slow), shadows, font stack (Inter, JetBrains Mono)", 1.5),
        ]
    },
    {
        "name": "2. ROOM SYSTEM & REAL-TIME INFRASTRUCTURE",
        "items": [
            ("Socket.IO event architecture", "29 typed events across client/server interfaces, room state broadcast, connection lifecycle management", 4),
            ("Room creation & management", "6-character alphanumeric room codes (collision-safe via nanoid), room state persistence, cleanup on empty", 3),
            ("Host role system", "Automatic host assignment on creation, host reassignment when host disconnects, crown badge UI indicator", 2),
            ("Connection resilience", "Auto-retry logic (3 retries + 15s timeout), connect_error handler, graceful redirect to home on failure", 3),
            ("User presence tracking", "Real-time user list with randomly assigned emoji avatars (16 pool), join/leave broadcasts, 'You' indicator", 2),
            ("Home / landing page", "Display name input (localStorage persist), create room flow, join by code with validation, animated gradient background", 3),
        ]
    },
    {
        "name": "3. VIDEO SYNCHRONIZATION ENGINE",
        "items": [
            ("YouTube IFrame Player API integration", "react-youtube wrapper component, player state management, event binding (play, pause, buffer, end, rate change)", 4),
            ("Sequence-based state sync", "seq number on every state update to prevent stale updates from being applied out of order across clients", 4),
            ("Latency-compensated time sync", "Elapsed-time compensation: targetTime = currentTime + network_elapsed_ms, 1.5s seek threshold to avoid unnecessary buffering", 5),
            ("Event deduplication & guards", "Debounced play/pause (150ms), isRemoteUpdate guard flag preventing echo loops, polling-based seek detection (1s interval, 2s threshold)", 5),
            ("Playback rate synchronization", "video:rate event broadcast, rate change detection and application across all connected clients", 2),
            ("VideoPlayer component (182 lines)", "Full player integration with all sync logic, state guards, and event handlers", 3),
        ]
    },
    {
        "name": "4. VIDEO QUEUE SYSTEM",
        "items": [
            ("Queue management logic", "Add to queue / play now dual-mode, reorder up/down, remove items, play specific item, play next (skip current)", 4),
            ("Auto-advance on video end", "Race condition protection via server-side endedProcessing Set with 2-second lock, queue splice and load", 3),
            ("YouTube metadata fetching", "Async oEmbed API title resolution (shows videoId immediately, updates when resolved), thumbnail previews", 2.5),
            ("QueuePanel component (108 lines)", "Drag-friendly reorder UI, thumbnail previews, 50-item cap, system chat messages for all queue actions", 3),
        ]
    },
    {
        "name": "5. REAL-TIME CHAT SYSTEM",
        "items": [
            ("Chat message system", "Socket.IO broadcast, system messages (join/leave/video load with distinct styling), 200-message rolling history", 3),
            ("Chat UX features", "Unread badge counter (clears on tab focus), auto-scroll to newest, 500-char limit, GIF URL detection with inline rendering", 3),
            ("Chat component (155 lines)", "Full chat interface with message input, history display, system message styling, GIF inline preview", 3),
        ]
    },
    {
        "name": "6. EMOJI & GIF PICKER INTEGRATION",
        "items": [
            ("Emoji picker integration", "@emoji-mart/react with dark theme, skin tone support, frequent row, click-outside-to-close behavior", 2.5),
            ("Giphy API integration", "Trending GIFs on open, search with 400ms debounce, masonry 2-column grid, lazy loading, GIPHY attribution", 4),
            ("GifPicker component (126 lines)", "Search input, grid layout, loading states, click-to-send, proper API key management", 2),
        ]
    },
    {
        "name": "7. YOUTUBE COMMENTS PANEL",
        "items": [
            ("Server-side Invidious proxy", "Proxy to open-source YouTube frontend API, 5 fallback instances tried in order for resilience", 4),
            ("Server-side caching layer", "5-minute comments cache (Map with expiry), auto-prune at 200 entries, keyed by videoId + sort", 2.5),
            ("Comments UI & pagination", "Sort by Top/Newest toggle, continuation token pagination, 'Load more' button, like counts with icons", 3),
            ("Reply thread system", "Collapsible reply threads with independent pagination, nested continuation tokens", 3),
            ("Content sanitization", "Unicode/invisible character stripping on author names (RTL/LTR marks, zero-width chars), author thumbnail with initial fallback", 1.5),
        ]
    },
    {
        "name": "8. WEBRTC VOICE CHAT (VOIP)",
        "items": [
            ("VoiceManager core class (638 lines)", "Peer connection Map management, joinVoice/leaveVoice lifecycle, AudioContext creation/resume, getUserMedia with constraints", 6),
            ("WebRTC signaling via Socket.IO", "Offer/answer/ICE candidate relay through server, 5 new server handlers, voice:active-users sync on join", 4),
            ("STUN/TURN server support", "Dynamic ICE config endpoint (/api/ice-servers), TURN env vars, 2x Google STUN + configurable TURN relay", 3),
            ("Web Audio API graph per peer", "MediaStreamSource -> GainNode -> AnalyserNode -> AudioContext.destination, separate input gain for mic volume", 4),
            ("Input device management", "Device selection dropdown, hot-swap via replaceTrack on all active peers, devicechange event listener for hot-plug", 2.5),
            ("Push-to-Talk system", "Configurable key binding, keydown/keyup handlers, window blur safety (releases key on focus loss), input field exclusion", 3),
            ("Voice Activity Detection (VAD)", "30fps AnalyserNode frequency polling for local + all remote peers, >15 amplitude threshold, 300ms debounce", 3),
            ("Mic level meter (2 variants)", "Horizontal bar (settings) + segmented dots (controls), requestAnimationFrame without React re-renders, green/yellow/red thresholds", 2.5),
            ("Error handling & edge cases", "DOMException parsing (NotAllowedError, NotFoundError), partial state cleanup, AudioContext suspended fix, double audio fix", 3),
            ("VoiceContext + VoiceControls UI", "React context wrapping VoiceManager (142 lines), floating controls overlay (80 lines), settings integration, localStorage persistence", 3),
        ]
    },
    {
        "name": "9. UI/UX DESIGN & THEMING",
        "items": [
            ("Glassmorphism design system", "backdrop-blur-xl panels, translucent backgrounds (white/[0.03]-[0.08]), subtle borders, layered depth effects", 5),
            ("6-theme color system", "Crimson, Ocean Blue, Neon Purple, Emerald, Sunset Orange, Midnight — CSS custom properties injected at runtime on documentElement", 4),
            ("Ambient gradient orbs", "Three large blurred background circles per theme (--orb-primary/secondary/tertiary), positioned at corners/center, theme-reactive", 2.5),
            ("Panel opacity system", "Settings slider (0%-20%), persisted to localStorage, applied as --panel-opacity CSS variable across all panels", 1.5),
            ("Custom CSS & animations (130 lines)", "Custom scrollbar styling, range input theming, fadeIn/slideUp keyframes, glow shadows, font imports", 3),
            ("ThemeContext provider (55 lines)", "Theme state management, localStorage persistence, CSS variable injection, theme list export", 2),
        ]
    },
    {
        "name": "10. RESPONSIVE LAYOUT & NAVIGATION",
        "items": [
            ("Desktop layout", "Fixed 380px sidebar with 5-tab navigation (Chat/People/Queue/Comments/Settings), video area fills remaining space", 3),
            ("Mobile layout", "Bottom navigation bar, slide-up overlay panel on tap, backdrop blur overlay with tap-to-close", 4),
            ("Tab system with badges", "User count badge, queue count badge, unread chat counter, active tab highlighting with accent color", 2),
            ("Room header component (87 lines)", "Room code display with copy-to-clipboard, compact mic mute toggle, leave room button with confirmation", 2),
        ]
    },
    {
        "name": "11. DEPLOYMENT & DEVOPS",
        "items": [
            ("Cloudflare Tunnel integration", "cloudflared service in Docker Compose, zero port forwarding, outbound-only tunnel, replaces Nginx Proxy Manager", 4),
            ("TURN server documentation", "Free hosted option (metered.ca) + self-hosted coturn with Docker Compose snippet and turnserver.conf config", 2),
            ("DEPLOYMENT.md guide (192 lines)", "Step-by-step: Cloudflare Zero Trust setup, env vars, TURN config, deploy commands, traffic flow diagram, NPM migration, troubleshooting table", 4),
            ("Health monitoring", "/api/health endpoint returning room count, user count, uptime; Docker healthcheck with depends_on condition", 1.5),
        ]
    },
    {
        "name": "12. TESTING, DEBUGGING & QA",
        "items": [
            ("Browser polyfill fix", "White screen debugging — simple-peer requires Buffer/process globals; installed vite-plugin-node-polyfills, configured Vite", 2),
            ("Voice over internet debugging", "Diagnosed missing TURN server for symmetric NAT traversal, implemented dynamic ICE config endpoint", 3),
            ("getUserMedia permission handling", "Diagnosed silent failure without user gesture, removed auto-join, added comprehensive error feedback UI", 2.5),
            ("Audio echo elimination", "Found dual playback (HTML audio + Web Audio API destination), muted audioEl when Web Audio succeeds", 2),
            ("Cross-browser audio context fix", "Diagnosed Chrome AudioContext suspended state, added resume() at creation, stream arrival, and user click events", 1.5),
            ("Voice user state tracking fix", "Fixed UserList showing no mic icons for silent users — added separate voiceUsers Set vs speakingUsers", 1.5),
            ("Connection & sync bug fixes", "Socket connection timing, stale state prevention, event storm debugging, race condition identification", 3),
            ("End-to-end scenario testing", "Multi-tab and multi-device testing, network condition simulation, edge case verification", 3),
        ]
    },
    {
        "name": "13. VOICE CHAT: P2P MESH → mediasoup SFU MIGRATION",
        "items": [
            ("Architecture design & migration planning", "Designed SFU topology replacing O(N²) P2P mesh with O(N) server-forwarded audio; mapped signaling protocol, transport lifecycle, and backward-compatible migration path preserving all existing audio processing", 4),
            ("mediasoup server infrastructure (mediasoup.ts — 357 lines, new file)", "Worker pool management, Router-per-room with lazy creation, WebRtcServer single-port multiplexing, WebRtcTransport factory (send/recv per peer), Producer/Consumer lifecycle, peer cleanup with producer-closed notifications, room cleanup on empty", 8),
            ("Socket.IO signaling protocol rewrite (index.ts)", "Replaced 3 P2P relay handlers (offer/answer/ICE) with 9 mediasoup negotiation handlers: voice:join (returns rtpCapabilities + existingProducers), create-send/recv-transport, connect-transport, produce, consume, resume-consumer, pause/resume-producer; updated handleDisconnect for SFU cleanup", 6),
            ("TypeScript event type definitions (types.ts)", "Removed P2P relay event types, defined 9 new ClientToServerEvents with typed callback signatures (transport params, consumer params, DTLS parameters), 2 new ServerToClientEvents (voice:new-producer, voice:producer-closed)", 2.5),
            ("VoiceManager complete rewrite (VoiceManager.ts — 664 lines changed)", "Replaced SimplePeer mesh with mediasoup-client Device/Transport/Producer/Consumer; preserved entire Web Audio pipeline (HighPass→LowPass→RNNoise→Compressor→NoiseGate→VAD); new consumeProducer/destroyConsumer flow; producer pause/resume for server-side mute optimization", 10),
            ("mediasoup-client transport negotiation", "Implemented createSendTransport/createRecvTransport with connect and produce event handlers, DTLS parameter exchange, connectionstatechange monitoring, ICE server injection for candidate gathering", 4),
            ("Cross-browser audio playback compatibility", "Diagnosed Chrome requiring Audio element for WebRTC RTP pipeline activation vs Firefox blocking Audio element autoplay outside user gestures; implemented hybrid approach: silent Audio element (volume=0) + AudioContext.destination for actual playback", 5),
            ("ICE connectivity debugging (Windows)", "Diagnosed and resolved ICE failures across 8+ iterations: loopback IP rejection by browsers, migrated from per-transport random ports to WebRtcServer single-port, auto-detect LAN IP via os.networkInterfaces(), Windows Firewall rules for port + mediasoup-worker.exe binary", 7),
            ("Docker & deployment configuration", "Updated Dockerfile with C++ build toolchain (python3, make, g++, linux-headers) for mediasoup worker compilation, exposed UDP/TCP port range 40000-40200, added MEDIASOUP_ANNOUNCED_IP/PORT/NUM_WORKERS env vars to docker-compose.yml", 3),
            ("Dependency integration & build fixes", "Installed mediasoup@3 (server) + mediasoup-client@3 (frontend), resolved TypeScript type issues (RouterRtpCodecCapability import path, implicit any on transport handlers, mediasoup-client re-exports), retained simple-peer for ScreenShare coexistence", 3.5),
            ("Multi-browser integration testing", "End-to-end testing across Firefox + Edge + Chrome, multi-tab verification, bidirectional audio flow confirmation, mute/unmute/PTT validation, speaking indicator (VAD) verification, quality preset switching, per-user volume control", 5),
        ]
    },
]

# ══════════════════════════════════════════════════════════════
# RENDER LINE ITEMS
# ══════════════════════════════════════════════════════════════
r = 11
total_hours = 0
section_fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

for sec in sections:
    # Section header
    set_row_height(r, 28)
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = section_fill
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    cell = ws.cell(row=r, column=2, value=sec["name"])
    cell.font = section_font
    cell.alignment = Alignment(vertical="center")
    r += 1

    section_hours = 0
    for idx, (task, desc, hours) in enumerate(sec["items"]):
        amount = hours * RATE
        section_hours += hours
        total_hours += hours

        bg = light_fill if idx % 2 == 0 else white_fill
        set_row_height(r, 38)

        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = bg
            ws.cell(row=r, column=c).border = thin_border

        # Task name
        ws.cell(row=r, column=2, value=task).font = item_font
        ws.cell(row=r, column=2).alignment = Alignment(vertical="center", wrap_text=True)

        # Description
        ws.cell(row=r, column=3, value=desc).font = item_desc_font
        ws.cell(row=r, column=3).alignment = Alignment(vertical="center", wrap_text=True)

        # Hours
        ws.cell(row=r, column=4, value=hours).font = item_font
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4).number_format = '0.0'

        # Rate
        ws.cell(row=r, column=5, value=RATE).font = item_font
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=5).number_format = '"$"#,##0.00'

        # Amount
        ws.cell(row=r, column=6, value=amount).font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=6).number_format = '"$"#,##0.00'

        r += 1

    # Section subtotal
    section_amount = section_hours * RATE
    set_row_height(r, 26)
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = subtotal_fill
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value=f"Subtotal — {sec['name'].split('. ', 1)[1]}").font = subtotal_font
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center")
    ws.cell(row=r, column=4, value=section_hours).font = subtotal_font
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=4).number_format = '0.0'
    ws.cell(row=r, column=6, value=section_amount).font = Font(name="Calibri", size=10, bold=True, color=ACCENT)
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
    r += 1

    # Spacer
    set_row_height(r, 6)
    r += 1

# ══════════════════════════════════════════════════════════════
# TOTALS
# ══════════════════════════════════════════════════════════════
total_amount = total_hours * RATE

# Blank row
set_row_height(r, 10)
r += 1

# Grand total bar
total_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
set_row_height(r, 50)
for c in range(1, 8):
    ws.cell(row=r, column=c).fill = total_fill

ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.cell(row=r, column=2, value="TOTAL").font = total_font
ws.cell(row=r, column=2).alignment = Alignment(vertical="center")

ws.cell(row=r, column=4, value=total_hours).font = Font(name="Calibri", size=14, bold=True, color=WHITE)
ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=r, column=4).number_format = '0.0'

ws.cell(row=r, column=5, value="hours x $250").font = Font(name="Calibri", size=10, color=ACCENT_LIGHT)
ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")

ws.cell(row=r, column=6, value=total_amount).font = total_amount_font
ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=r, column=6).number_format = '"$"#,##0.00'

r += 1

# ══════════════════════════════════════════════════════════════
# BREAKDOWN SUMMARY
# ══════════════════════════════════════════════════════════════
set_row_height(r, 15)
r += 1

set_row_height(r, 24)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws.cell(row=r, column=2, value="HOURS BREAKDOWN BY PHASE").font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
ws.cell(row=r, column=2).alignment = Alignment(vertical="center")
ws.cell(row=r, column=2).border = Border(bottom=Side(style="medium", color=ACCENT))
for c in range(3, 7):
    ws.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=ACCENT))
r += 1

phase_data = []
for sec in sections:
    sec_hours = sum(h for _, _, h in sec["items"])
    phase_data.append((sec["name"], sec_hours))

for name, hrs in phase_data:
    set_row_height(r, 22)
    pct = (hrs / total_hours) * 100
    bar = "█" * int(pct / 2.5) + "░" * (40 - int(pct / 2.5))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(row=r, column=2, value=name).font = Font(name="Calibri", size=9, color=TEXT_DARK)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center")
    ws.cell(row=r, column=4, value=f"{hrs:.1f} hrs").font = Font(name="Calibri", size=9, bold=True, color=TEXT_DARK)
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=5, value=f"{pct:.0f}%").font = Font(name="Calibri", size=9, color=TEXT_MED)
    ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=6, value=f"${hrs * RATE:,.2f}").font = Font(name="Calibri", size=9, bold=True, color=TEXT_DARK)
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=r, column=2).border = thin_border
    for c in range(3, 7):
        ws.cell(row=r, column=c).border = thin_border
    r += 1

# ══════════════════════════════════════════════════════════════
# NOTES / TERMS
# ══════════════════════════════════════════════════════════════
r += 1
set_row_height(r, 24)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws.cell(row=r, column=2, value="PROJECT DELIVERABLES").font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
ws.cell(row=r, column=2).border = Border(bottom=Side(style="medium", color=ACCENT))
for c in range(3, 7):
    ws.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=ACCENT))
r += 1

deliverables = [
    "Complete source code: React/TypeScript frontend + Node.js/TypeScript server (~4,800 lines)",
    "Docker Compose deployment stack with Cloudflare Tunnel (zero port forwarding)",
    "Real-time synchronized YouTube video player with sequence-based state sync",
    "mediasoup SFU voice chat supporting 20+ concurrent users (migrated from P2P mesh)",
    "Real-time chat with emoji picker, GIF integration (Giphy API), and inline media",
    "Video queue system with auto-advance, reordering, and metadata fetching",
    "YouTube comments panel via Invidious proxy with caching and pagination",
    "Glassmorphism UI with 6 color themes, panel opacity control, and responsive layout",
    "Comprehensive deployment documentation (DEPLOYMENT.md)",
]

for d in deliverables:
    set_row_height(r, 18)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2, value=f"  •  {d}").font = notes_font
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center")
    r += 1

r += 1
set_row_height(r, 24)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
ws.cell(row=r, column=2, value="TERMS").font = Font(name="Calibri", size=11, bold=True, color=ACCENT)
ws.cell(row=r, column=2).border = Border(bottom=Side(style="medium", color=ACCENT))
for c in range(3, 7):
    ws.cell(row=r, column=c).border = Border(bottom=Side(style="medium", color=ACCENT))
r += 1

terms = [
    "Payment due within 30 days of invoice date",
    "All source code and intellectual property transferred to client upon payment",
    "Includes source code, Docker configuration, and deployment documentation",
    "Hours include development, architecture design, debugging, testing, and documentation",
]

for t in terms:
    set_row_height(r, 18)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2, value=f"  •  {t}").font = notes_font
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center")
    r += 1

# ══════════════════════════════════════════════════════════════
# PRINT SETTINGS
# ══════════════════════════════════════════════════════════════
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = "landscape"
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3

# Freeze panes at header
ws.freeze_panes = "A11"

# Save
output_path = r"C:\Users\demon\personal\WatchTogether\WatchTogether_Invoice.xlsx"
wb.save(output_path)
print(f"Invoice generated: {output_path}")
print(f"Total hours: {total_hours}")
print(f"Total amount: ${total_amount:,.2f}")
